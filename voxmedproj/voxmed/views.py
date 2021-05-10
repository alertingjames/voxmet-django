import datetime
import difflib
import os
import string
import urllib
from itertools import islice

import io
import requests
import xlrd
import re

from django.core import mail
from django.core.mail import send_mail, BadHeaderError, EmailMessage
from django.contrib import messages
# from _mysql_exceptions import DataError, IntegrityError
from django.template import RequestContext

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives

from django.core.files.storage import FileSystemStorage
import json
from django.contrib import auth
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect, HttpResponseNotAllowed
from django.utils.datastructures import MultiValueDictKeyError
from django.views.decorators.cache import cache_control
from numpy import long

import pandas as pd
import numpy as np

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.fields import empty
from rest_framework.permissions import AllowAny
from xlrd import XLRDError
from time import gmtime, strftime
import time
from openpyxl.styles import PatternFill

from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.contrib.sessions.backends.db import SessionStore
from django.contrib.sessions.models import Session
from django.contrib.auth.models import User, AnonymousUser
from django.conf import settings
from django import forms
import sys
from django.core.cache import cache

import urllib.request
import urllib.parse
from random import randint
import random
import math
import wave

from voxmed.models import Member, Report, Account, ReportPicture, Field, Template, Keyword
from voxmed.serializers import MemberSerializer, ReportSerializer, ReportPictureSerializer, FieldSerializer, TemplateSerializer, KeywordSerializer


def index(request):
    return HttpResponse('<h2>Hello I am Voxmed backend!</h2>')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def registerMember(request):

    if request.method == 'POST':

        name = request.POST.get('name', '')
        eml = request.POST.get('email', '')
        password = request.POST.get('password', '')
        age = request.POST.get('age', '')
        role = request.POST.get('role', '')
        patientID = request.POST.get('patientID', '')

        users = Member.objects.filter(email=eml)
        count = users.count()
        if count == 0:
            member = Member()
            member.name = name
            member.email = eml
            member.picture_url = settings.URL + '/static/voxmed/images/anonymous.png'
            member.password = password
            if age != '': member.age = age
            member.role = role
            member.patientID = patientID
            member.registered_time = str(int(round(time.time() * 1000)))
            member.save()

            fs = FileSystemStorage()

            i = 0
            for f in request.FILES.getlist('files'):
                # print("Product File Size: " + str(f.size))
                # if f.size > 1024 * 1024 * 2:
                #     continue
                i = i + 1
                filename = fs.save(f.name, f)
                uploaded_url = fs.url(filename)
                if i == 1:
                    member.picture_url = settings.URL + uploaded_url
                    member.save()

            serializer = MemberSerializer(member, many=False)

            resp = {'result_code': '0', 'data':serializer.data}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:
            users = Member.objects.filter(email=eml, role=role)
            count = users.count()
            if count == 0:
                resp_er = {'result_code': '1'}
                return HttpResponse(json.dumps(resp_er))
            else:
                resp_er = {'result_code': '2'}
                return HttpResponse(json.dumps(resp_er))

    elif request.method == 'GET':
        pass


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def login(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')
        role = request.POST.get('role', '')
        if password != '':
            members = Member.objects.filter(email=email, password=password, role=role)
        else:
            members = Member.objects.filter(email=email, role=role)
        resp = {}
        if members.count() > 0:
            member = members[0]
            serializer = MemberSerializer(member, many=False)
            resp = {'result_code': '0', 'data':serializer.data}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)
        else:
            members = Member.objects.filter(email=email, role=role)
            if members.count() > 0:
                resp = {'result_code': '2'}
            else: resp = {'result_code':'1'}
        return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def forgotpassword(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')

        usrs = Member.objects.filter(email=email)
        if usrs.count() == 0:
            return HttpResponse(json.dumps({'result_code': '1'}))

        message = 'You are allowed to reset your password from your request.<br>For it, please click this link to reset your password.<br><br>https://voxmed.pythonanywhere.com/resetpassword?email=' + email

        html =  """\
                    <html>
                        <head></head>
                        <body>
                            <a href="https://voxmed.pythonanywhere.com/"><img src="https://voxmed.pythonanywhere.com/static/voxmed/images/logo.png" style="width:150px;height:150px;border-radius: 8%; margin-left:25px;"/></a>
                            <h2 style="margin-left:10px; color:#02839a;">Voxmed User's Security Update Information</h2>
                            <div style="font-size:16px; word-break: break-all; word-wrap: break-word;">
                                {mes}
                            </div>
                        </body>
                    </html>
                """
        html = html.format(mes=message)

        fromEmail = 'voxmedro@gmail.com'
        toEmailList = []
        toEmailList.append(email)
        msg = EmailMultiAlternatives('We allowed you to reset your password', '', fromEmail, toEmailList)
        msg.attach_alternative(html, "text/html")
        msg.send(fail_silently=False)

        return HttpResponse(json.dumps({'result_code': '0'}))

def resetpassword(request):
    email = request.GET['email']
    return render(request, 'voxmed/resetpwd.html', {'email':email})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rstpwd(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')
        repassword = request.POST.get('repassword', '')
        if len(password) < 8:
            return render(request, 'voxmed/result.html',
                          {'response': 'Please enter password of characters more than 8.'})
        if password != repassword:
            return render(request, 'voxmed/result.html',
                          {'response': 'Please enter the same password.'})
        members = Member.objects.filter(email=email)
        if members.count() > 0:
            member = members[0]
            member.password = password
            member.save()
            return render(request, 'voxmed/result.html',
                          {'response': 'Password has been reset successfully.'})
        else:
            return render(request, 'voxmed/result.html',
                          {'response': 'You haven\'t been registered.'})
    else: pass


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def postreport(request):
    if request.method == 'POST':
        patientID = request.POST.get('patientID', '1')
        member_id = request.POST.get('member_id', '1')
        report_body = request.POST.get('report', '')
        sts = request.POST.get('status', '')
        temp_json_str = request.POST.get('temp_json_str', '')

        members = Member.objects.filter(patientID=patientID)
        if members.count() == 0:
            resp = {'result_code':'1'}
            return HttpResponse(json.dumps(resp))

        report = Report()
        report.patientID = patientID
        report.member_id = member_id
        report.body = report_body
        report.date_time = str(int(round(time.time() * 1000)))
        report.status = sts

        report.save()

        fs = FileSystemStorage()

        i = 0
        for f in request.FILES.getlist('files'):
            # print("Product File Size: " + str(f.size))
            # if f.size > 1024 * 1024 * 2:
            #     continue
            i = i + 1
            filename = fs.save(f.name, f)
            uploaded_url = fs.url(filename)
            if i == 1:
                report.audio_url = settings.URL + uploaded_url
                report.save()
            else:
                if report.picture_url == '':
                    report.picture_url = settings.URL + uploaded_url
                    report.save()
                p = ReportPicture()
                p.report_id = report.pk
                p.picture_url = settings.URL + uploaded_url
                p.save()


        try:
            decoded = json.loads(temp_json_str)
            for field_data in decoded['fields']:

                title = field_data['title']
                content = field_data['content']

                field = Field()
                field.member_id = member_id
                field.report_id = report.pk
                field.title = title
                field.content = content

                field.save()

            resp = {'result_code':'0'}
            return HttpResponse(json.dumps(resp))

        except:
            resp = {'result_code': '2'}
            return HttpResponse(json.dumps(resp))

        resp = {'result_code':'0'}
        return HttpResponse(json.dumps(resp))

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getallusers(request):
    if request.method == 'POST':
        members = Member.objects.all().order_by('-id')
        serializer = MemberSerializer(members, many=True)
        resp = {'result_code':'0', 'data':serializer.data}
        return HttpResponse(json.dumps(resp))

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getallreports(request):
    if request.method == 'POST':
        reports = Report.objects.all().order_by('-id')
        serializer = ReportSerializer(reports, many=True)
        resp = {'result_code':'0', 'data':serializer.data}
        return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getreportpictures(request):
    if request.method == 'POST':
        report_id = request.POST.get('report_id', '1')
        pics = ReportPicture.objects.filter(report_id=report_id)
        serializer = ReportPictureSerializer(pics, many=True)
        resp = {'result_code':'0', 'data':serializer.data}
        return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getmember(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', '1')
        members = Member.objects.filter(id=member_id)
        if members.count() > 0:
            member = members[0]
            serializer = MemberSerializer(member, many=False)
            resp = {'result_code':'0', 'data':serializer.data}
            return HttpResponse(json.dumps(resp))
        else:
            resp = {'result_code':'1'}
            return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updatereport(request):
    if request.method == 'POST':
        report_id = request.POST.get('report_id', '1')
        report_body = request.POST.get('report', '')
        sts = request.POST.get('status', '')
        temp_json_str = request.POST.get('temp_json_str', '')

        reports = Report.objects.filter(id=int(report_id))
        if reports.count() == 0:
            resp = {'result_code':'1'}
            return HttpResponse(json.dumps(resp))

        report = reports[0]
        report.body = report_body
        report.date_time = str(int(round(time.time() * 1000)))
        if report.status == '':
            report.status = sts

        report.save()

        try:
            decoded = json.loads(temp_json_str)

            for field_data in decoded['fields']:
                fid = field_data['id']
                title = field_data['title']
                content = field_data['content']

                field = Field.objects.get(id=fid)
                field.title = title
                field.content = content

                field.save()

            resp = {'result_code':'0'}
            return HttpResponse(json.dumps(resp))

        except:
            resp = {'result_code': '2'}
            return HttpResponse(json.dumps(resp))

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp))

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updatemember(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', '1')
        name = request.POST.get('name', '')
        age = request.POST.get('age', '')

        members = Member.objects.filter(id=member_id)
        if members.count() == 0:
            resp = {'result_code':'1'}
            return HttpResponse(json.dumps(resp))

        member = members[0]
        member.name = name
        if age != '': member.age = age

        member.save()

        fs = FileSystemStorage()

        i = 0
        for f in request.FILES.getlist('files'):
            # print("Product File Size: " + str(f.size))
            # if f.size > 1024 * 1024 * 2:
            #     continue
            i = i + 1
            filename = fs.save(f.name, f)
            uploaded_url = fs.url(filename)
            if i == 1:
                member.picture_url = settings.URL + uploaded_url
                member.save()

        serializer = MemberSerializer(member, many=False)
        resp = {'result_code': '0', 'data':serializer.data}

        return HttpResponse(json.dumps(resp))


def newpic(request):
    pic = ReportPicture()
    pic.save()
    return HttpResponse('success')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getReportFields(request):
    if request.method == 'POST':
        report_id = request.POST.get('report_id', '1')
        fields = Field.objects.filter(report_id=report_id)
        serializer = FieldSerializer(fields, many=True)
        resp = {'result_code':'0', 'data':serializer.data}

        return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def saveKeywords(request):
    if request.method == 'POST':
        temp_id = request.POST.get('temp_id', '0')
        member_id = request.POST.get('member_id', '1')
        name = request.POST.get('name', '')
        keyword_json_str = request.POST.get('keyword_json_str', '')

        members = Member.objects.filter(id=member_id)
        if members.count() == 0:
            resp = {'result_code':'1'}
            return HttpResponse(json.dumps(resp))

        temp = None
        if int(temp_id) == 0:
            temp = Template()
        else:
            temps = Template.objects.filter(id=temp_id)
            if temps.count() > 0: temp = temps[0]

        if temp is not None:
            temp.member_id = member_id
            temp.name = name
            temp.save()
        try:
            decoded = json.loads(keyword_json_str)

            if int(temp_id) > 0:
                keywords = Keyword.objects.filter(template_id=temp_id)
                for keyword in keywords: keyword.delete()

            for json_data in decoded['keywords']:

                kwd = json_data['keyword']

                keyword = Keyword()
                keyword.template_id = temp.pk
                keyword.keyword = kwd
                keyword.save()

            keywords = Keyword.objects.filter(template_id=temp.pk)
            temp.items_count = str(keywords.count())
            temp.save()

            resp = {'result_code':'0'}
            return HttpResponse(json.dumps(resp))

        except:
            resp = {'result_code': '2'}
            return HttpResponse(json.dumps(resp))

        return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getKeywords(request):
    if request.method == 'POST':
        temp_id = request.POST.get('temp_id', '1')
        kwds = Keyword.objects.filter(template_id=temp_id)
        serializer = KeywordSerializer(kwds, many=True)
        resp = {'result_code':'0', 'data':serializer.data}

        return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getTemplates(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', '1')
        temps = Template.objects.filter(member_id=member_id).order_by('-id')
        serializer = TemplateSerializer(temps, many=True)
        resp = {'result_code':'0', 'data':serializer.data}

        return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def deleteTemplate(request):
    if request.method == 'POST':
        temp_id = request.POST.get('temp_id', '1')
        temps = Template.objects.filter(id=temp_id)
        if temps.count() > 0:
            temp = temps[0]
            keywords = Keyword.objects.filter(template_id=temp.pk)
            for keyword in keywords: keyword.delete()
            temp.delete()
            resp = {'result_code':'0'}
        else: resp = {'result_code':'1'}
        return HttpResponse(json.dumps(resp))



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def mailReport(request):
    if request.method == 'POST':
        to_email = request.POST.get('to_email', '')
        from_email = request.POST.get('from_email', '')
        temp_json_str = request.POST.get('temp_json_str', '')

        members = Member.objects.filter(email=from_email)
        if members.count() > 0: member = members[0]

        usrs = Member.objects.filter(email=to_email)
        if usrs.count() == 0:
            return HttpResponse(json.dumps({'result_code': '1'}))

        message = ''

        try:
            decoded = json.loads(temp_json_str)

            for field_data in decoded['fields']:

                title = field_data['title']
                content = field_data['content']

                space = ''
                if message == '': space = ''
                else: space = '<br>'
                message = message + space + "<label style='font-weight:600; font-size:16px;'>" + title + ": " + "</label>" + content

        except:
            resp = {'result_code': '2'}
            return HttpResponse(json.dumps(resp))

        message = message + '<br><br>' + member.name

        html =  """\
                    <html>
                        <head></head>
                        <body>
                            <img src="https://voxmed.pythonanywhere.com/static/voxmed/images/logo.png" style="width:150px;height:150px;border-radius: 8%; margin-left:25px;"/>
                            <h2 style="margin-left:10px; color:#02839a;">Voxmed Medical Report</h2>
                            <div style="font-size:16px; word-break: break-all; word-wrap: break-word;">
                                {mes}
                            </div>
                        </body>
                    </html>
                """
        html = html.format(mes=message)

        fromEmail = from_email
        toEmailList = []
        toEmailList.append(to_email)
        msg = EmailMultiAlternatives('Voxmed Report Sharing', '', fromEmail, toEmailList)
        msg.attach_alternative(html, "text/html")
        msg.send(fail_silently=False)

        return HttpResponse(json.dumps({'result_code': '0'}))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getAllKeywords(request):
    if request.method == 'POST':
        member_id = request.POST.get('member_id', '1')
        temps = Template.objects.filter(member_id=member_id)
        keywordList = []
        keyList = []
        for temp in temps:
            kwds = Keyword.objects.filter(template_id=temp.pk)
            for kwd in kwds:
                if kwd.keyword not in keyList:
                    keyList.append(kwd.keyword)
                    keywordList.append(kwd)
        serializer = KeywordSerializer(keywordList, many=True)
        resp = {'result_code':'0', 'data':serializer.data}

        return HttpResponse(json.dumps(resp))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def getpatient(request):
    if request.method == 'POST':
        patientID = request.POST.get('patientID', '')
        patients = Member.objects.filter(patientID=patientID)
        if patients.count() == 0:
            resp = {'result_code':'1'}
            return HttpResponse(json.dumps(resp))
        patient = patients[0]
        serializer = MemberSerializer(patient, many=False)

        resp = {'result_code':'0', 'data':serializer.data}
        return HttpResponse(json.dumps(resp))















































########################## Website #############################################################################################################################################################################

def loginpage(request):
    try:
        if request.session['status'] != '':
            return redirect('/home')
    except KeyError:
        print('no session')
    return render(request, 'voxmed/login.html')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def adminloginprocess(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')
        role = request.POST.get('role', '')
        if role == 'admin':
            accounts = Account.objects.filter(email=email, password=password)
            if accounts.count() > 0:
                members = Member.objects.filter(email=email, password=password)
                member = None
                if members.count() == 0:
                    member = Member()
                    member.name = 'Administrator'
                    member.email = email
                    member.password = password
                    member.role = 'admin'
                    member.save()
                else:
                    member = members[0]
                if member is not None: request.session['ID'] = member.pk
                reports = Report.objects.all().order_by('-id')
                return render(request, 'voxmed/home.html', {'reports':getFullReportData(reports), 'member':member})
            else:
                # account = Account()
                # account.save()
                return render(request, 'voxmed/result.html',
                              {'response': 'You don\'t have any permission to access this site. Try again with another credential.'})
        else:
            members = Member.objects.filter(email=email, password=password, role=role)
            if members.count() > 0:
                member = members[0]
                request.session['ID'] = member.pk
                if role == 'doctor':
                    reports = Report.objects.filter(member_id=member.pk).order_by('-id')
                elif role == 'employee':
                    reports = Report.objects.all().order_by('-id')
                return render(request, 'voxmed/home.html', {'reports':getFullReportData(reports), 'member':member})
            else:
                return render(request, 'voxmed/result.html',
                              {'response': 'You haven\'t been registered in this site. Please signup with Voxmed app as a member.'})
    elif request.method == 'GET':
        return redirect('/home')
    else:
        return redirect('/home')


def logout(request):
    request.session['ID'] = 0
    return render(request, 'voxmed/login.html')


def home(request):
    if request.session['ID'] == 0:
        return redirect('/logout')
    members = Member.objects.filter(id=request.session['ID'])
    if members.count() > 0:
        member = members[0]
        if member.role == 'doctor':
            reports = Report.objects.filter(member_id=member.pk).order_by('-id')
        elif member.role == 'employee':
            reports = Report.objects.all().order_by('-id')
        elif member.role == 'admin':
            reports = Report.objects.all().order_by('-id')
        return render(request, 'voxmed/home.html', {'reports':getFullReportData(reports), 'member':member})
    else: return redirect('/logout')


def getFullReportData(reports):
    reportList = []
    for report in reports:
        doctor = None
        patient = None
        doctors = Member.objects.filter(id=report.member_id)
        if doctors.count() > 0:
            doctor = doctors[0]
        patients = Member.objects.filter(patientID=report.patientID)
        if patients.count() > 0:
            patient = patients[0]
        report_body = ''
        fields = Field.objects.filter(report_id=report.pk)
        for field in fields:
            space = ''
            if report_body == '': space = ''
            else: space = '; '
            report_body = report_body + space + field.title + ': ' + field.content
        data = {
            'id': report.pk,
            'doctor':doctor,
            'patient':patient,
            'report':report_body,
            'audio':report.audio_url,
            'picture':report.picture_url,
            'date_time':report.date_time,
            'status':report.status
        }
        reportList.append(data)
    return reportList


def reports(request):
    if request.session['ID'] == 0:
        return redirect('/logout')
    member_id = request.GET['member_id']
    if member_id is not None:
        members = Member.objects.filter(id=member_id)
        if members.count() > 0:
            member = members[0]
            if member.role == 'doctor':
                reports = Report.objects.filter(member_id=member.pk).order_by('-id')
            elif member.role == 'patient':
                reports = Report.objects.filter(patientID=member.patientID).order_by('-id')
            return render(request, 'voxmed/reports.html', {'reports':getFullReportData(reports), 'member':member})
    else: return redirect('/logout')



def doctors(request):
    if request.session['ID'] == 0:
        return redirect('/logout')
    members = Member.objects.filter(id=request.session['ID'])
    if members.count() > 0:
        member = members[0]
        if member.role == 'admin':
            doctors = Member.objects.filter(role='doctor').order_by('-id')
        return render(request, 'voxmed/doctors.html', {'doctors':getDoctorData(doctors), 'member':member})
    else: return redirect('/logout')

def getDoctorData(doctors):
    doctorList = []
    for doctor in doctors:
        reports = Report.objects.filter(member_id=doctor.pk)
        data = {
            'id': doctor.pk,
            'picture':doctor.picture_url,
            'name':doctor.name,
            'email':doctor.email,
            'registered_time':doctor.registered_time,
            'reports':reports.count()
        }
        doctorList.append(data)
    return doctorList

def employees(request):
    if request.session['ID'] == 0:
        return redirect('/logout')
    members = Member.objects.filter(id=request.session['ID'])
    if members.count() > 0:
        member = members[0]
        if member.role == 'admin':
            employees = Member.objects.filter(role='employee').order_by('-id')
        return render(request, 'voxmed/employees.html', {'employees':employees, 'member':member})
    else: return redirect('/logout')


def patients(request):
    if request.session['ID'] == 0:
        return redirect('/logout')
    members = Member.objects.filter(id=request.session['ID'])
    if members.count() > 0:
        member = members[0]
        if member.role == 'admin':
            patients = Member.objects.filter(role='patient').order_by('-id')
        return render(request, 'voxmed/patients.html', {'patients':getPatientData(patients), 'member':member})
    else: return redirect('/logout')


def getPatientData(patients):
    patientList = []
    for patient in patients:
        reports = Report.objects.filter(patientID=patient.patientID)
        data = {
            'id': patient.pk,
            'picture':patient.picture_url,
            'name':patient.name,
            'email':patient.email,
            'age':patient.age,
            'registered_time':patient.registered_time,
            'reports':reports.count()
        }
        patientList.append(data)
    return patientList


def adminsetting(request):
    if request.session['ID'] == 0:
        return redirect('/logout')
    members = Member.objects.filter(id=request.session['ID'])
    if members.count() > 0:
        member = members[0]
        accounts = Account.objects.all()
        return  render(request, 'voxmed/account.html', {'member':member, 'accounts':accounts})
    else: return redirect('/logout')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def addadminaccount(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')
        accounts = Account.objects.filter(email=email)
        if accounts.count() > 0:
            return render(request, 'voxmed/result.html',
                          {'response': 'You already are using the same email. Try again with another email.'})
        account = Account()
        account.email = email
        account.password = password
        account.save()

        return redirect('/adminsetting')

def deleteAccount(request, account_id):
    account = Account.objects.get(id=account_id)
    account.delete()
    return redirect('/adminsetting')


def removemember(request):
    member_id = request.GET['member_id']
    members = Member.objects.filter(id=member_id)
    if members.count() == 0:
        return redirect('/logout')

    member = members[0]
    role = member.role

    temps = Template.objects.filter(member_id=member.pk)
    for temp in temps:
        keywords = Keyword.objects.filter(template_id=temp.pk)
        for keyword in keywords:
            keyword.delete()
        temp.delete()

    fs = FileSystemStorage()

    reports = Report.objects.filter(member_id=member.pk)
    for report in reports:

        fname = report.picture_url.replace(settings.URL + '/media/', '')
        fs.delete(fname)
        fname = report.audio_url.replace(settings.URL + '/media/', '')
        fs.delete(fname)

        fields = Field.objects.filter(report_id=report.pk)
        for field in fields: field.delete()

        report.delete()

    if role == 'patient':
        reports = Report.objects.filter(patientID=member.patientID)
        for report in reports:

            fname = report.picture_url.replace(settings.URL + '/media/', '')
            fs.delete(fname)
            fname = report.audio_url.replace(settings.URL + '/media/', '')
            fs.delete(fname)

            fields = Field.objects.filter(report_id=report.pk)
            for field in fields: field.delete()

            report.delete()

    fname = member.picture_url.replace(settings.URL + '/media/', '')
    fs.delete(fname)
    member.delete()

    if role == 'doctor':
        return redirect('/doctors')
    elif role == 'employee':
        return redirect('/employees')
    elif role == 'patient':
        return redirect('/patients')
    else:
        return redirect('/home')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def newTemplateAdd(request):
    if request.method == 'POST':
        name = request.POST.get('name', '')
        if request.session['ID'] == 0:
            return redirect('/logout')
        members = Member.objects.filter(id=request.session['ID'])
        if members.count() > 0:
            member = members[0]
            temps = Template.objects.filter(name=name, member_id=member.pk)
            if temps.count() > 0:
                return render(request, 'voxmed/result.html',
                          {'response': 'The template name already exists.'})
            else:
                temp = Template()
                temp.member_id = member.pk
                temp.name = name
                temp.items_count = '0'
                temp.save()
                return redirect('/newreport')
        else:
            return redirect('/logout')



def manage(request):
    if request.session['ID'] == 0:
        return redirect('/logout')
    members = Member.objects.filter(id=request.session['ID'])
    if members.count() > 0:
        member = members[0]
        report_id = request.GET['report_id']
        reports = Report.objects.filter(id=report_id)
        if reports.count() > 0:
            report = reports[0]

            fields = Field.objects.filter(report_id=report.pk)

            doctor = None
            patient = None
            doctors = Member.objects.filter(id=report.member_id)
            if doctors.count() > 0:
                doctor = doctors[0]
            patients = Member.objects.filter(patientID=report.patientID)
            if patients.count() > 0:
                patient = patients[0]
            return render(request, 'voxmed/manage_report.html', {'report':report, 'fields':fields, 'member':member, 'doctor':doctor, 'patient':patient})
        else:
            return redirect('/home')
    else: return redirect('/logout')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def reportupdate(request):
    if request.method == 'POST':
        report_id = request.POST.get('report_id', '0')
        option = request.POST.get('option', '')
        fieldids = request.POST.getlist('fieldid[]')
        contents = request.POST.getlist('body[]')

        reports = Report.objects.filter(id=report_id)
        if reports.count() == 0:
            return render(request, 'voxmed/result.html',
                          {'response': 'The report doesn\'t exist.'})

        report = reports[0]
        message = ''
        for id in fieldids:

            index = fieldids.index(id)
            content = contents[index]

            if index == len(fieldids) - 1:
                report.body = content
                report.save()

            field = Field.objects.get(id=id)
            field.report_id = report.pk
            field.content = content

            field.save()

            title = field.title
            content = field.content

            space = ''
            if message == '': space = ''
            else: space = '<br>'
            message = message + space + "<label style='font-weight:600; font-size:16px;'>" + title + ": " + "</label>" + content

        if option == 'share':
            to_email = request.POST.get('to_email', '')

            return HttpResponse(to_email)

            members = Member.objects.filter(id=request.session['ID'])
            if members.count() > 0:
                member = members[0]

            usrs = Member.objects.filter(email=to_email)
            if usrs.count() == 0:
                return render(request, 'voxmed/result.html',
                              {'response': 'The receiver doesn\'t exist.'})

            message = message + '<br><br>' + member.name

            html =  """\
                        <html>
                            <head></head>
                            <body>
                                <img src="https://voxmed.pythonanywhere.com/static/voxmed/images/logo.png" style="width:150px;height:150px;border-radius: 8%; margin-left:25px;"/>
                                <h2 style="margin-left:10px; color:#02839a;">Voxmed Medical Report</h2>
                                <div style="font-size:16px; word-break: break-all; word-wrap: break-word;">
                                    {mes}
                                </div>
                            </body>
                        </html>
                    """
            html = html.format(mes=message)

            fromEmail = member.email
            toEmailList = []
            toEmailList.append(to_email)
            msg = EmailMultiAlternatives('Voxmed Report Sharing', '', fromEmail, toEmailList)
            msg.attach_alternative(html, "text/html")
            msg.send(fail_silently=False)

            return render(request, 'voxmed/result.html',
                              {'response': 'The report has been shared via email.'})


        elif option == '':
            return render(request, 'voxmed/result.html',
                          {'response': 'The report has been updated.'})
        elif option == 'export':
            return redirect('/export_xlsx?report_id=' + report_id)


def export_xlsx(request):

    report_id = request.GET['report_id']
    report = Report.objects.get(id=report_id)

    fields = Field.objects.filter(report_id=report.pk)

    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    import urllib3
    from requests.packages.urllib3.exceptions import MaxRetryError

    docName = "Voxmed Report"

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + docName + '.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Report"

    row_num = 0

    columns = [
        (u"Keyword", 30),
        (u"Content", 100),
    ]

    my_color = openpyxl.styles.colors.Color(rgb='FF99EDFC')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)

    from openpyxl.styles import Alignment

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]
        c.fill = my_fill
        c.alignment = Alignment(horizontal='center')

    row_num = 1

    row_num = row_num + 1

    c = ws.cell(row=row_num, column=1)
    c.value = 'PatientID'

    c = ws.cell(row=row_num, column=2)
    c.value = report.patientID

    row_num = row_num + 1

    c = ws.cell(row=row_num, column=1)
    c.value = 'Patient Name'

    c = ws.cell(row=row_num, column=2)
    patients = Member.objects.filter(patientID=report.patientID)
    patient = patients[0]
    c.value = patient.name

    for field in fields:
        row_num = row_num + 1

        c = ws.cell(row=row_num, column=1)
        c.value = field.title

        c = ws.cell(row=row_num, column=2)
        c.value = field.content

    ws2 = wb.create_sheet(title='Pictures')
    row_num = 0

    columns = [
        (u"Picture", 100),
    ]

    for col_num in range(len(columns)):
        c = ws2.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws2.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]
        c.fill = my_fill
        c.alignment = Alignment(horizontal='center')

    row_num = 1
    pics = ReportPicture.objects.filter(report_id=report.pk)

    for pic in pics:
        row_num = row_num + 1

        c = ws2.cell(row=row_num, column=1)
        c.value = pic.picture_url

    ws3 = wb.create_sheet(title='Audio')
    row_num = 0

    columns = [
        (u"Audio_URL", 100),
    ]

    for col_num in range(len(columns)):
        c = ws3.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws3.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]
        c.fill = my_fill
        c.alignment = Alignment(horizontal='center')

    c = ws3.cell(row=2, column=1)
    c.value = report.audio_url

    wb.save(response)
    return response



def pictures(request):
    report_id = request.GET['report_id']
    report = Report.objects.get(id=report_id)
    pictures = ReportPicture.objects.filter(report_id=report_id)

    return render(request, 'voxmed/gallery.html', {'pictures':pictures, 'report':report})


def newreport(request):

#    return render(request, 'voxmed/test.html')          ################################################

    if request.session['ID'] == 0:
        return redirect('/logout')
    members = Member.objects.filter(id=request.session['ID'])
    if members.count() > 0:
        member = members[0]
        temps = Template.objects.filter(member_id=member.pk).order_by('-id')
        tempList = []
        for temp in temps:
            keywords = Keyword.objects.filter(template_id=temp.pk)
            data = {
                'id':temp.pk,
                'member_id':member.pk,
                'name':temp.name,
                'items_count': str(keywords.count()),
                'keywords':keywords
            }
            tempList.append(data)
        return render(request, 'voxmed/templates.html', {'templates':tempList, 'member':member})
    else: return redirect('/logout')


def tempdelete(request):
    temp_id = request.GET['temp_id']
    temps = Template.objects.filter(id=temp_id)
    if temps.count() > 0:
        temp = temps[0]
        temp.delete()
        keywords = Keyword.objects.filter(template_id=temp.pk)
        for keyword in keywords:
            keyword.delete()
    return redirect('/newreport')


def keyworddelete(request):
    k_id = request.GET['keyword_id']
    keywords = Keyword.objects.filter(id=k_id)
    if keywords.count() > 0:
        keyword = keywords[0]
        keyword.delete()
    return redirect('/newreport')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def newKeywordAdd(request):
    if request.method == 'POST':
        temp_id = request.POST.get('temp_id', '1')
        template = Template.objects.get(id=temp_id)
        keywordstr = request.POST.get('new_keyword', '')
        keywords = Keyword.objects.filter(template_id=temp_id, keyword=keywordstr)
        if keywords.count() == 0:
            keyword = Keyword()
            keyword.template_id = temp_id
            keyword.keyword = keywordstr
            keyword.save()

            template.items_count = str(int(template.items_count) + 1)
            template.save()

            return redirect('/newreport')
        else:
            return render(request, 'voxmed/result.html',
                          {'response': 'The keyword already exists in the template.'})


def tempchoose(request):
    temp_id = request.GET['temp_id']
    if request.session['ID'] == 0:
        return redirect('/logout')
    members = Member.objects.filter(id=request.session['ID'])
    if members.count() > 0:
        member = members[0]
    kwds = Keyword.objects.filter(template_id=temp_id)
    patients = Member.objects.filter(role='patient').order_by('-id')

    temps = Template.objects.filter(member_id=member.pk)
    keywordList = []
    keyList = []
    for temp in temps:
        mykwds = Keyword.objects.filter(template_id=temp.pk)
        for kwd in mykwds:
            if kwd.keyword not in keyList:
                keyList.append(kwd.keyword)
                keywordList.append(kwd)

    return render(request, 'voxmed/new_report.html', {'keywords':kwds, 'allkeywords':keywordList, 'patients':getPatientData(patients), 'member':member})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def reportpost(request):
    if request.method == 'POST':
        patient_id = request.POST.get('patient_id', '0')
        members = Member.objects.filter(id=request.session['ID'])
        if members.count() > 0:
            member = members[0]
        keyword_contents = request.POST.getlist('body[]')
        keyword_titles = request.POST.getlist('keyword[]')
        sts = request.POST.get('status', '')
        audio = request.FILES['audio']

        patients = Member.objects.filter(id=patient_id)
        if patients.count() == 0:
            return render(request, 'voxmed/result.html',
                          {'response': 'The patient doesn\'t exist.'})

        fs = FileSystemStorage()
        filename = fs.save(audio.name, audio)
        uploaded_url = fs.url(filename)

        report = Report()
        patient = patients[0]
        report.patientID = patient.patientID
        report.member_id = member.pk
        report.audio_url = settings.URL + uploaded_url
        report.date_time = str(int(round(time.time() * 1000)))
        report.status = sts

        report.save()

        for f in request.FILES.getlist('pictures'):
            # print("Product File Size: " + str(f.size))
            # if f.size > 1024 * 1024 * 2:
            #     continue
            filename = fs.save(f.name, f)
            uploaded_url = fs.url(filename)
            if report.picture_url == '':
                report.picture_url = settings.URL + uploaded_url
                report.save()
            p = ReportPicture()
            p.report_id = report.pk
            p.picture_url = settings.URL + uploaded_url
            p.save()

        for title in keyword_titles:

            index = keyword_titles.index(title)
            content = keyword_contents[index]

            if index == len(keyword_titles) - 1:
                report.body = content
                report.save()

            field = Field()
            field.member_id = member.pk
            field.report_id = report.pk
            field.title = title
            field.content = content

            field.save()

        return redirect('/home')

def reportdelete(request):
    report_id = request.POST.get('report_id', '0')
    report = Report.objects.get(id=report_id)
    fs = FileSystemStorage()
    rps = ReportPicture.objects.filter(report_id=report.pk)
    for rp in rps:
        fname = rp.picture_url.replace(settings.URL + '/media/', '')
        fs.delete(fname)
        rp.delete()
    fields = Field.objects.filter(report_id=report.pk)
    for field in fields:
        field.delete()
    report.delete()
    return redirect('/home')


def docpatients(request):
    if request.session['ID'] == 0:
        return redirect('/logout')
    members = Member.objects.filter(id=request.session['ID'])
    if members.count() > 0:
        member = members[0]
        if member.role == 'doctor':
            patients = Member.objects.filter(role='patient').order_by('-id')
        return render(request, 'voxmed/docpatients.html', {'patients':getPatientData(patients), 'member':member})
    else: return redirect('/logout')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def newpatient(request):

    if request.method == 'POST':

        name = request.POST.get('name', '')
        eml = request.POST.get('email', '')
        age = request.POST.get('age', '')
        patient_id = request.POST.get('patient_id', '0')

        users = Member.objects.filter(email=eml)
        count = users.count()
        if count == 0:
            if int(patient_id) == 0:
                member = Member()
                member.name = name
                member.email = eml
                member.picture_url = settings.URL + '/static/voxmed/images/anonymous.png'
                if age != '': member.age = age
                member.role = 'patient'
                res = ''.join(random.choices(string.ascii_uppercase +
                                 string.digits, k = 10))
                member.patientID = str(res)
                member.registered_time = str(int(round(time.time() * 1000)))
                member.save()

                fs = FileSystemStorage()

                try:
                    f = request.FILES['picture']
                    filename = fs.save(f.name, f)
                    uploaded_url = fs.url(filename)
                    member.picture_url = settings.URL + uploaded_url
                    member.save()
                except MultiValueDictKeyError:
                    print('No exists')

                return redirect('/docpatients')

        else:
            if int(patient_id) > 0:
                member = Member.objects.get(id=patient_id)
                member.name = name
                member.email = eml
                if age != '': member.age = age
                member.role = 'patient'
                member.save()

                fs = FileSystemStorage()

                try:
                    f = request.FILES['picture']
                    filename = fs.save(f.name, f)
                    uploaded_url = fs.url(filename)
                    member.picture_url = settings.URL + uploaded_url
                    member.save()
                except MultiValueDictKeyError:
                    print('No exists')

                return redirect('/docpatients')

    elif request.method == 'GET':
        pass



def patientremove(request):
    member_id = request.GET['member_id']
    members = Member.objects.filter(id=member_id)
    if members.count() == 0:
        return redirect('/logout')

    member = members[0]
    role = member.role

    temps = Template.objects.filter(member_id=member.pk)
    for temp in temps:
        keywords = Keyword.objects.filter(template_id=temp.pk)
        for keyword in keywords:
            keyword.delete()
        temp.delete()

    fs = FileSystemStorage()

    reports = Report.objects.filter(member_id=member.pk)
    for report in reports:

        fname = report.picture_url.replace(settings.URL + '/media/', '')
        fs.delete(fname)
        fname = report.audio_url.replace(settings.URL + '/media/', '')
        fs.delete(fname)

        fields = Field.objects.filter(report_id=report.pk)
        for field in fields: field.delete()

        report.delete()

    if role == 'patient':
        reports = Report.objects.filter(patientID=member.patientID)
        for report in reports:

            fname = report.picture_url.replace(settings.URL + '/media/', '')
            fs.delete(fname)
            fname = report.audio_url.replace(settings.URL + '/media/', '')
            fs.delete(fname)

            fields = Field.objects.filter(report_id=report.pk)
            for field in fields: field.delete()

            report.delete()

    fname = member.picture_url.replace(settings.URL + '/media/', '')
    fs.delete(fname)
    member.delete()

    return redirect('/docpatients')
















































































def test(request):
    from urllib.request import urlopen
    pcmfile = urlopen(settings.URL + "/media/recording.pcm")
    with open('recording.pcm','wb') as output:
      output.write(pcmfile.read())

    with wave.open('test.wav', 'wb') as wavfile:
        wavfile.setparams((2, 2, 44100, 0, 'NONE', 'NONE'))
        wavfile.writeframes(bytes(pcmfile))

        time.sleep(60)

        response = HttpResponse(wavfile, content_type="application/waveform")
        response['Content-Disposition'] = 'inline; filename=' + os.path.basename('test.wav')
        return response

    # convert_dir(r'/media/', '.pcm')
    # return HttpResponse('success')


def is_wav(f):
    res = True
    try:
        wave.open(f)
    except wave.Error as e:
        res = False
    return res


def pcm2wav(pcm_file, save_file, channels=1, bits=64, sample_rate=44100):
    """ pcm转换为wav格式
        Args:
            pcm_file pcm文件
            save_file 保存文件
            channels 通道数
            bits 量化位数，即每个采样点占用的比特数
            sample_rate 采样频率
    """
    if is_wav(pcm_file):
        raise ValueError('"' + str(pcm_file) + '"' +
                         " is a wav file, not pcm file! ")

    pcmf = open(pcm_file, 'rb')
    pcmdata = pcmf.read()
    pcmf.close()

    if bits % 8 != 0:
        raise ValueError("bits % 8 must == 0. now bits:" + str(bits))

    wavfile = wave.open(save_file, 'wb')

    wavfile.setnchannels(channels)
    wavfile.setsampwidth(bits // 8)
    wavfile.setframerate(sample_rate)

    wavfile.writeframes(pcmdata)
    wavfile.close()


def convert_dir(root, ext=".pcm", **kwargs):
    """ 把一个文件夹内的pcm，统统加上头
        Args:
            root 文件夹根目录
            ext pcm文件的扩展名
    """
    from tqdm import tqdm
    src_files = [os.path.join(dir_path, f)
                 for dir_path, _, files in os.walk(root)
                 for f in files
                 if os.path.splitext(f)[1] == ext]

    for src_file in tqdm(src_files, ascii=True):
        try:
            wav_file = os.path.splitext(src_file)[0] + ".wav"
            pcm2wav(src_file, wav_file, **kwargs)
        except Exception as e:
            print('Convert fail: ' + src_file)
            print(e)

































